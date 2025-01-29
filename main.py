import streamlit as st
import pandas as pd
import uuid
import datetime
from openpyxl import Workbook, load_workbook

# Define the menu items with prices
menu = {
    "01": {"name": "Chicken Fried Rice", "price": 80},
    "02": {"name": "Veg Fried Rice", "price": 70},
    "03": {"name": "Chicken Noodles", "price": 90},
    "04": {"name": "Veg Noodles", "price": 75}
}

# Excel file for storing billing data
excel_file = "billing_data.xlsx"

# Initialize or load the Excel file
def initialize_excel():
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        if ws.max_row == 1:  # If only headers exist, reinitialize
            ws.append(["Date", "Bill Number", "Phone Number", "Item Name", "Quantity", "Price", "Total Amount", "Status"])
            wb.save(excel_file)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Billing Data"
        ws.append(["Date", "Bill Number", "Phone Number", "Item Name", "Quantity", "Price", "Total Amount", "Status"])
        wb.save(excel_file)

initialize_excel()

# Streamlit UI Setup
st.set_page_config(page_title="The Taste Quest", page_icon="🍔", layout="wide")

# Title in the center
st.markdown("<h1 style='text-align: center;'>🍔 The Taste Quest</h1>", unsafe_allow_html=True)

# Layout - Two Columns
col1, col2 = st.columns([1, 1])  # Left for customer details & history, right for billing

# Left Side (Customer Phone + Billing History)
with col1:
    st.subheader("📞 Customer Details")
    phone_number = st.text_input("Enter Customer Phone Number:")

    # Generate unique bill number (same bill number for multiple items in one session)
    if "bill_number" not in st.session_state:
        st.session_state["bill_number"] = str(uuid.uuid4())[:8]
    bill_number = st.session_state["bill_number"]

    # Show past billing history for entered phone number
    if phone_number:
        try:
            df = pd.read_excel(excel_file)  # Read all columns
            df.columns = ["Date", "Bill Number", "Phone Number", "Item Name", "Quantity", "Price", "Total Amount", "Status"]
            customer_history = df[df["Phone Number"] == phone_number]

            if not customer_history.empty:
                customer_history["Date"] = pd.to_datetime(customer_history["Date"])  # Convert to datetime
                customer_history = customer_history.sort_values(by="Date", ascending=False)  # Sort by date

                st.subheader("📜 Past Orders")
                st.dataframe(customer_history)
            else:
                st.write("No previous records found for this number.")
        except Exception as e:
            st.error(f"Error loading billing history: {e}")

# Right Side (Item Selection + Quantity + Add to Bill)
with col2:
    st.subheader("🛒 Order Details")

    # Auto-complete for item selection
    item_names = {v["name"]: k for k, v in menu.items()}  # Reverse lookup for codes
    selected_item = st.selectbox("🔍 Select an item:", options=list(item_names.keys()))

    # Display price in a small card
    if selected_item:
        item_code = item_names[selected_item]
        item = menu[item_code]
        st.markdown(f"**Price:** Rs. {item['price']}")

    quantity = st.number_input("📦 Enter quantity:", min_value=1, step=1)

    if st.button("✅ Add to Bill"):
        if selected_item and phone_number:
            item_code = item_names[selected_item]
            item = menu[item_code]
            total_amount = item["price"] * quantity

            # Get the current date & time
            current_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Save to Excel properly
            wb = load_workbook(excel_file)
            ws = wb.active
            ws.append([
                current_date,       # Date
                str(bill_number),   # Bill Number
                str(phone_number),  # Phone Number
                item["name"],       # Item Name
                int(quantity),      # Quantity
                int(item["price"]), # Price
                int(total_amount),  # Total Amount
                "Pending"           # Status
            ])
            wb.save(excel_file)

            st.success(f"✅ {item['name']} added! Total Amount: Rs. {total_amount}")

        else:
            st.error("❌ Please enter a phone number and select an item.")

# Pending Orders Section at the bottom right
st.subheader("📝 Pending Orders")

try:
    df = pd.read_excel(excel_file)  # Read all columns
    df.columns = ["Date", "Bill Number", "Phone Number", "Item Name", "Quantity", "Price", "Total Amount", "Status"]
    pending_orders = df[df["Status"] == "Pending"]

    if not pending_orders.empty:
        for index, row in pending_orders.iterrows():
            col1, col2 = st.columns([3, 1])
            with col1:
                st.write(f"**Bill Number:** {row['Bill Number']} | **Item:** {row['Item Name']} | **Quantity:** {row['Quantity']} | **Total Amount:** Rs. {row['Total Amount']}")
            with col2:
                if st.checkbox("Mark as Completed", key=index):
                    # Update status to "Completed"
                    wb = load_workbook(excel_file)
                    ws = wb.active
                    ws.cell(row=index + 2, column=8, value="Completed")  # Update status in Excel
                    wb.save(excel_file)
                    st.rerun()  # Refresh the page to reflect changes
    else:
        st.write("No pending orders.")
except Exception as e:
    st.error(f"Error loading pending orders: {e}")